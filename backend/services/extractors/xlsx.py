"""Извлечение из XLSX/XLS (ТЗ 4.2). openpyxl (xlsx) + xlrd (старый xls).

Обходим все листы книги, определяем строку заголовков (может быть не первой),
ищем колонки названия услуги и цен (резидент/нерезидент).
"""
import logging

from services.extractors import ExtractResult, RawRow
from services.extractors.row_parser import to_number, classify_columns, detect_currency

logger = logging.getLogger(__name__)


def extract(file_path: str) -> ExtractResult:
    result = ExtractResult()
    try:
        sheets = _load_sheets(file_path)   # [(title, rows[][]), ...]
    except Exception as e:  # noqa: BLE001
        logger.exception('xls/xlsx load failed')
        result.warnings.append(f'Ошибка чтения файла Excel: {e}')
        return result

    for title, rows in sheets:
        if not rows:
            continue
        header_idx, col_map = classify_columns(rows)
        if col_map.get('name') is None:
            result.warnings.append(f'Лист "{title}": не найдена колонка услуги')
            continue
        # текст заголовка — подсказка валюты, если она задана в шапке колонки
        # («Цена, USD»), а не в каждой ячейке
        header_text = ' '.join(str(c) for c in rows[header_idx] if c is not None)
        for r in rows[header_idx + 1:]:
            name = r[col_map['name']] if col_map['name'] < len(r) else None
            if not name or not str(name).strip():
                continue
            res_cell = _cell(r, col_map.get('price_resident'))
            nonres_cell = _cell(r, col_map.get('price_nonresident'))
            result.raw_text += ' | '.join(str(c) for c in r if c is not None) + '\n'
            result.rows.append(RawRow(
                service_name_raw=str(name).strip(),
                price_resident=to_number(res_cell),
                price_nonresident=to_number(nonres_cell),
                service_code_source=_str(_cell(r, col_map.get('code'))),
                currency=detect_currency([res_cell, nonres_cell, header_text]),
            ))
    return result


def _load_sheets(file_path: str):
    """Вернуть [(title, rows)] для книги. .xls читаем через xlrd, остальное — openpyxl."""
    if file_path.lower().endswith('.xls'):
        return _load_xls(file_path)
    return _load_xlsx(file_path)


def _load_xlsx(file_path: str):
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True, read_only=True)
    return [(ws.title, [list(row) for row in ws.iter_rows(values_only=True)])
            for ws in wb.worksheets]


def _load_xls(file_path: str):
    """Старый бинарный формат .xls (openpyxl его не читает) — через xlrd."""
    import xlrd
    book = xlrd.open_workbook(file_path)
    out = []
    for sheet in book.sheets():
        rows = [sheet.row_values(r) for r in range(sheet.nrows)]
        out.append((sheet.name, rows))
    return out


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _str(v):
    return str(v).strip() if v is not None else None
