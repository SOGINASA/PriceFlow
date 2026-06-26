"""Загрузка целевого справочника услуг (ТЗ 2.2 / 4.3).

Справочник даётся организаторами в момент старта в XLSX или JSON.
Здесь — приём, upsert в таблицу Service и (опционально) переразметка позиций.
"""
import logging

from flask import Blueprint, request, jsonify

from models import db, Service, PriceItem
from services import normalization_service as norm
from services import catalog_service
from services.extractors.row_parser import to_number  # noqa: F401 (consistency)

logger = logging.getLogger(__name__)
catalog_bp = Blueprint('catalog', __name__)


@catalog_bp.route('/build', methods=['POST'])
def build_catalog():
    """Сформировать справочник из загруженных прайсов (ТЗ §7) и перепривязать позиции.

    Body (опционально): {threshold: 0..100, only_unmatched: bool}.
    """
    data = request.get_json(silent=True) or {}
    summary = catalog_service.build_catalog_from_items(
        threshold=int(data.get('threshold', 90)),
        only_unmatched=data.get('only_unmatched', True),
    )
    return jsonify(summary)


@catalog_bp.route('/export', methods=['GET'])
def export_catalog():
    """Выгрузить текущий справочник (JSON, совместим с /catalog/import)."""
    resp = jsonify(catalog_service.export_catalog())
    resp.headers['Content-Disposition'] = 'attachment; filename=services_catalog.json'
    return resp


@catalog_bp.route('/consolidate', methods=['POST'])
def consolidate_catalog():
    """Свести дубликаты-синонимы справочника к каноническому названию (ТЗ 4.3).

    Body (опционально): {use_llm: bool}. По умолчанию use_llm=true (LLM-канонизация
    при наличии ключа GROQ_API_KEY; иначе офлайн-нормализация).
    """
    data = request.get_json(silent=True) or {}
    summary = catalog_service.consolidate_catalog(use_llm=data.get('use_llm', True))
    return jsonify(summary)


@catalog_bp.route('/import', methods=['POST'])
def import_catalog():
    """Импорт справочника. Принимает JSON-массив услуг или XLSX-файл (multipart).

    JSON-элемент: {service_id?, service_name, synonyms[], category, icd_code}
    После upsert справочника автоматически пересопоставляет уже загруженные
    несопоставленные позиции (service_id IS NULL). Отключается через ?rematch=0.
    """
    created, updated = 0, 0

    if request.files.get('file'):
        rows = _parse_xlsx(request.files['file'])
    else:
        rows = request.get_json(silent=True) or []

    for r in rows:
        name = (r.get('service_name') or '').strip()
        if not name:
            continue
        svc = None
        if r.get('service_id'):
            svc = db.session.get(Service, r['service_id'])
        if not svc:
            svc = Service.query.filter(db.func.lower(Service.service_name) == name.lower()).first()
        if svc:
            updated += 1
        else:
            svc = Service(service_id=r.get('service_id'))
            db.session.add(svc)
            created += 1
        svc.service_name = name
        svc.synonyms = r.get('synonyms') or []
        svc.category = r.get('category')
        svc.icd_code = r.get('icd_code')
        svc.is_active = r.get('is_active', True)

    db.session.commit()

    rematched = 0
    if request.args.get('rematch') != '0':
        rematched = _rematch_unmatched()

    return jsonify({'created': created, 'updated': updated,
                    'total': created + updated, 'rematched': rematched})


def _rematch_unmatched():
    """Прогнать автосопоставление по позициям без service_id (после смены справочника).

    Возвращает число позиций, которые удалось привязать к услуге справочника.
    """
    index = norm._build_index()
    if not index:
        return 0
    items = (PriceItem.query
             .filter(PriceItem.service_id.is_(None), PriceItem.is_active.is_(True))
             .all())
    matched = 0
    for item in items:
        if norm.normalize_item(item, index):
            matched += 1
    if items:
        db.session.commit()
    return matched


def _parse_xlsx(file_storage):
    """Разобрать справочник из XLSX: колонки service_name, synonyms, category, icd_code."""
    from openpyxl import load_workbook
    wb = load_workbook(file_storage, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip().lower() if c else '' for c in rows[0]]
    idx = {name: header.index(name) for name in
           ('service_name', 'synonyms', 'category', 'icd_code') if name in header}
    out = []
    for r in rows[1:]:
        def cell(key):
            return r[idx[key]] if key in idx and idx[key] < len(r) else None
        syn = cell('synonyms')
        out.append({
            'service_name': cell('service_name'),
            'synonyms': [s.strip() for s in str(syn).split(';') if s.strip()] if syn else [],
            'category': cell('category'),
            'icd_code': cell('icd_code'),
        })
    return out
